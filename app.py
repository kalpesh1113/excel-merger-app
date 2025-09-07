import streamlit as st
import pandas as pd
import io
import xlrd
import openpyxl
from openpyxl.styles import Border, Side
import datetime


st.set_page_config(page_title="Excel Merger Tool", layout="centered")

st.title("📊 Excel Merger Tool")

st.write("""
यह tool multiple Excel files को merge करता है:

1. हर file की शुरुआती 4 rows हटाता है (header merge वाले).
2. Row no. 5 को header मानता है.
3. Column **BU** के हिसाब से **TIME SLOT** column auto भरता है → `APP-BAL-<BU>`.
4. सभी files merge होकर एक ही Excel बनती है.
5. Final Excel में पूरी sheet को **All Borders** apply किए जाते हैं.
6. Output filename इस format में होगा: `4158_4341_4359_BalanceConsForReading_07Sep2025.xlsx`
""")


uploaded_files = st.file_uploader(
    "Select Excel Files (.xls / .xlsx)", type=["xls", "xlsx"], accept_multiple_files=True
)


# ---- Helper function to convert .xls → .xlsx ----
def convert_xls_to_xlsx(xls_bytes):
    book_xls = xlrd.open_workbook(file_contents=xls_bytes)
    book_xlsx = openpyxl.Workbook()
    sheet_xlsx = book_xlsx.active

    sheet_xls = book_xls.sheet_by_index(0)
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            sheet_xlsx.cell(row=row+1, column=col+1).value = sheet_xls.cell_value(row, col)

    # Save to bytes buffer
    xlsx_bytes = io.BytesIO()
    book_xlsx.save(xlsx_bytes)
    xlsx_bytes.seek(0)
    return xlsx_bytes


# ---- Helper function to add borders to all cells ----
def add_borders_to_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(file_path)


if st.button("🚀 Merge Files"):
    if not uploaded_files:
        st.warning("Please upload Excel files first.")
    else:
        merged_df = None
        try:
            for i, file in enumerate(uploaded_files):
                file_bytes = file.read()
                
                # Convert old .xls to .xlsx
                if file.name.endswith(".xls"):
                    st.info(f"Converting {file.name} to .xlsx format...")
                    file_bytes = convert_xls_to_xlsx(file_bytes).read()
                
                # Read into pandas
                df = pd.read_excel(io.BytesIO(file_bytes), skiprows=4, dtype=str)

                if i == 0:
                    merged_df = df.copy()
                else:
                    df = df.iloc[1:].reset_index(drop=True)
                    merged_df = pd.concat([merged_df, df], ignore_index=True)

            if merged_df is not None:
                # Normalize column names
                merged_df.columns = merged_df.columns.str.strip().str.upper()

                # BU → TIME SLOT mapping
                if "BU" in merged_df.columns and "TIME SLOT" in merged_df.columns:
                    merged_df["TIME SLOT"] = "APP-BAL-" + merged_df["BU"].astype(str)
                else:
                    st.warning("⚠️ 'BU' or 'TIME SLOT' column not found. Columns found: " + str(list(merged_df.columns)))

                # ✅ Unique BU values for filename
                if "BU" in merged_df.columns:
                    unique_bus = merged_df["BU"].dropna().unique().tolist()
                    bu_part = "_".join(unique_bus)
                else:
                    bu_part = "BU"

                # ✅ Date part for filename
                date_str = datetime.datetime.now().strftime("%d%b%Y")  # e.g. 07Sep2025

                # ✅ Final filename
                output_file = f"AppBalanceConsForReading_{date_str}.xlsx"

                merged_df.to_excel(output_file, index=False)

                # ✅ Add all borders
                add_borders_to_excel(output_file)

                # Download button
                with open(output_file, "rb") as f:
                    st.download_button(
                        label="📥 Download Merged Excel",
                        data=f,
                        file_name=output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                st.success(f"✅ Merged file created successfully: {output_file}")
        except Exception as e:
            st.error(f"Error: {e}")
